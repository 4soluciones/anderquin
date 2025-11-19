import decimal
import os

import reportlab
import io
from datetime import datetime, timedelta, date

import requests
from dateutil.relativedelta import relativedelta
from django.http import HttpResponse
from reportlab.lib.colors import Color, black, white, HexColor
from reportlab.lib.pagesizes import landscape, A5, portrait, A6, A4, letter
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle, Spacer, Image, Flowable
from reportlab.platypus import Table
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.barcode import qr
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY, TA_RIGHT
from reportlab.lib import colors
from reportlab.lib.units import cm, inch
from reportlab.rl_settings import defaultPageSize
# from reportlab.platypus.tables import ROUNDEDCORNERS

# Para Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from anderquin import settings
from .models import BillPurchase, Bill, CashFlow
from ..hrm.models import Worker, Subsidiary
from ..hrm.views import get_subsidiary_by_user
from ..buys.models import Purchase, PurchaseDetail, MoneyChange
from ..sales.models import Supplier, ProductSupplier, ProductDetail, Client, LoanPayment
from ..sales.number_to_letters import numero_a_moneda
from django.contrib.auth.models import User

PAGE_HEIGHT = defaultPageSize[1]
PAGE_WIDTH = defaultPageSize[0]

COLOR_PDF = colors.Color(red=(152.0 / 255), green=(29.0 / 255), blue=(31.0 / 255))
COLOR_BLUE = colors.Color(red=(27.0 / 255), green=(140.0 / 255), blue=(66.0 / 255))
COLOR_BLUE = colors.Color(red=(133.0 / 255), green=(180.0 / 255), blue=(242.0 / 255))

styles = getSampleStyleSheet()
styles.add(ParagraphStyle(name='Right', alignment=TA_RIGHT, leading=30, fontName='Square', fontSize=25))
styles.add(ParagraphStyle(name='JustifyTitle', alignment=TA_CENTER, leading=30, fontName='Square-Bold', fontSize=25))
styles.add(ParagraphStyle(name='Justify', alignment=TA_JUSTIFY, leading=8, fontName='Square', fontSize=8))
styles.add(ParagraphStyle(name='JustifySquare', alignment=TA_JUSTIFY, leading=12, fontName='Square', fontSize=8))
styles.add(ParagraphStyle(name='LeftSquare', alignment=TA_LEFT, leading=12, fontName='Square', fontSize=13))
styles.add(ParagraphStyle(name='LeftSquareSmall', alignment=TA_LEFT, leading=9, fontName='Square', fontSize=10))
styles.add(ParagraphStyle(name='LeftSquareSmall2', alignment=TA_LEFT, leading=9, fontName='Square', fontSize=8))
styles.add(ParagraphStyle(name='Justify-Dotcirful', alignment=TA_JUSTIFY, leading=12, fontName='Dotcirful-Regular',
                          fontSize=10))
styles.add(
    ParagraphStyle(name='Justify-Dotcirful-table', alignment=TA_JUSTIFY, leading=12, fontName='Dotcirful-Regular',
                   fontSize=7))
styles.add(ParagraphStyle(name='Justify_Bold', alignment=TA_JUSTIFY, leading=8, fontName='Square-Bold', fontSize=8))

styles.add(ParagraphStyle(name='Center', alignment=TA_CENTER, leading=50, fontName='Square-Bold', fontSize=35,
                          textColor=colors.black))
styles.add(ParagraphStyle(name='Center-fecha', alignment=TA_CENTER, leading=10, fontName='Square-Bold', fontSize=10,
                          textColor=colors.black))
styles.add(ParagraphStyle(name='Center-datos', alignment=TA_CENTER, leading=10, fontName='Square', fontSize=10,
                          textColor=colors.black))
styles.add(ParagraphStyle(name='Center-arequipa', alignment=TA_CENTER, leading=19, fontName='Square-Bold', fontSize=10,
                          textColor=COLOR_PDF))
# styles.add(ParagraphStyle(name='Center-titulo', alignment=TA_CENTER, leading=20, fontName='Square-Bold', fontSize=20,
#                          textColor=colors.steelblue))
styles.add(ParagraphStyle(name='Center-titulo', alignment=TA_CENTER, leading=40, fontName='Square-Bold', fontSize=40,
                          textColor=colors.black))
styles.add(ParagraphStyle(name='Center-recibo', alignment=TA_CENTER, leading=20, fontName='Square-Bold', fontSize=20,
                          textColor=colors.white))
styles.add(ParagraphStyle(name='Center-id', alignment=TA_CENTER, leading=40, fontName='Lucida-Console', fontSize=30,
                          textColor=colors.black))
styles.add(ParagraphStyle(name='Center-ng', alignment=TA_CENTER, leading=10, fontName='Square-Bold', fontSize=10,
                          textColor=colors.white))
styles.add(
    ParagraphStyle(name='Left', alignment=TA_LEFT, leading=30, fontName='Square', fontSize=25, textColor=colors.black))
styles.add(
    ParagraphStyle(name='CenterSquare', alignment=TA_CENTER, leading=30, fontName='Square', fontSize=25, textColor=colors.black))
styles.add(
    ParagraphStyle(name='Left-Simple', alignment=TA_LEFT, leading=15, fontName='Square', fontSize=15,
                   textColor=colors.black))
styles.add(ParagraphStyle(name='Left-name', alignment=TA_LEFT, leading=8, fontName='Square-Bold', fontSize=8,
                          textColor=COLOR_PDF))
styles.add(ParagraphStyle(name='Left-datos', alignment=TA_LEFT, leading=10, fontName='Square-Bold', fontSize=10,
                          textColor=colors.black))

styles.add(ParagraphStyle(name='Center4', alignment=TA_CENTER, leading=12, fontName='Square-Bold',
                          fontSize=14, spaceBefore=6, spaceAfter=6))
styles.add(ParagraphStyle(name='Center5', alignment=TA_LEFT, leading=15, fontName='ticketing.regular',
                          fontSize=12))
styles.add(
    ParagraphStyle(name='Center-Dotcirful', alignment=TA_CENTER, leading=12, fontName='Dotcirful-Regular', fontSize=10))
styles.add(ParagraphStyle(name='CenterTitle', alignment=TA_CENTER, leading=8, fontName='Square-Bold', fontSize=8))
styles.add(ParagraphStyle(name='CenterTitle-Dotcirful', alignment=TA_CENTER, leading=12, fontName='Dotcirful-Regular',
                          fontSize=10))
styles.add(ParagraphStyle(name='CenterTitle2', alignment=TA_CENTER, leading=8, fontName='Square-Bold', fontSize=12))
styles.add(ParagraphStyle(name='Center_Regular', alignment=TA_CENTER, leading=8, fontName='Ticketing', fontSize=10))
styles.add(ParagraphStyle(name='Center_Bold', alignment=TA_CENTER,
                          leading=8, fontName='Square-Bold', fontSize=12, spaceBefore=6, spaceAfter=6))
styles.add(ParagraphStyle(name='ticketing.regular', alignment=TA_CENTER,
                          leading=8, fontName='ticketing.regular', fontSize=14, spaceBefore=6, spaceAfter=6))
styles.add(ParagraphStyle(name='Center2', alignment=TA_CENTER, leading=8, fontName='Ticketing', fontSize=8))
styles.add(ParagraphStyle(name='Center3', alignment=TA_JUSTIFY, leading=8, fontName='Ticketing', fontSize=6))
styles.add(
    ParagraphStyle(name='Justify_Newgot_title', alignment=TA_JUSTIFY, leading=14, fontName='Newgot', fontSize=14))
styles.add(
    ParagraphStyle(name='Center_Newgot_title', alignment=TA_CENTER, leading=15, fontName='Newgot', fontSize=15))
styles.add(
    ParagraphStyle(name='Center_Newgot_title_small', alignment=TA_CENTER, leading=12, fontName='Newgot', fontSize=12))
styles.add(ParagraphStyle(name='Left_Square', alignment=TA_LEFT, leading=10, fontName='Square', fontSize=10))
styles.add(ParagraphStyle(name='Center_Square', alignment=TA_CENTER, leading=10, fontName='Square', fontSize=10))
styles.add(ParagraphStyle(name='Justify_Square', alignment=TA_JUSTIFY, leading=10, fontName='Square', fontSize=9))
styles.add(ParagraphStyle(name='Center_Newgot_1', alignment=TA_CENTER, leading=11, fontName='Newgot', fontSize=9))
styles.add(ParagraphStyle(name='Center-text', alignment=TA_CENTER, leading=8, fontName='Square', fontSize=8))
styles.add(ParagraphStyle(name='Justify_Newgot', alignment=TA_JUSTIFY, leading=10, fontName='Newgot', fontSize=10))
styles.add(ParagraphStyle(name='Right_Newgot', alignment=TA_RIGHT, leading=12, fontName='Newgot', fontSize=12))
style = styles["Normal"]

reportlab.rl_config.TTFSearchPath.append(str(settings.BASE_DIR) + '/static/fonts')
pdfmetrics.registerFont(TTFont('Narrow', 'Arial Narrow.ttf'))
pdfmetrics.registerFont(TTFont('Square', 'square-721-condensed-bt.ttf'))
pdfmetrics.registerFont(TTFont('Square-Bold', 'sqr721bc.ttf'))
pdfmetrics.registerFont(TTFont('Newgot', 'newgotbc.ttf'))
pdfmetrics.registerFont(TTFont('Ticketing', 'ticketing.regular.ttf'))
pdfmetrics.registerFont(TTFont('Lucida-Console', 'lucida-console.ttf'))
pdfmetrics.registerFont(TTFont('Square-Dot', 'square_dot_digital-7.ttf'))
pdfmetrics.registerFont(TTFont('Serif-Dot', 'serif_dot_digital-7.ttf'))
pdfmetrics.registerFont(TTFont('Enhanced-Dot-Digital', 'enhanced-dot-digital-7.regular.ttf'))
pdfmetrics.registerFont(TTFont('Merchant-Copy-Wide', 'MerchantCopyWide.ttf'))
pdfmetrics.registerFont(TTFont('Dot-Digital', 'dot_digital-7.ttf'))
pdfmetrics.registerFont(TTFont('Raleway-Dots-Regular', 'RalewayDotsRegular.ttf'))
pdfmetrics.registerFont(TTFont('Ordre-Depart', 'Ordre-de-Depart.ttf'))
pdfmetrics.registerFont(TTFont('Dotcirful-Regular', 'DotcirfulRegular.otf'))
pdfmetrics.registerFont(TTFont('Nationfd', 'Nationfd.ttf'))
pdfmetrics.registerFont(TTFont('Kg-Primary-Dots', 'KgPrimaryDots-Pl0E.ttf'))
pdfmetrics.registerFont(TTFont('Dot-line', 'Dotline-LA7g.ttf'))
pdfmetrics.registerFont(TTFont('Dot-line-Light', 'DotlineLight-XXeo.ttf'))
pdfmetrics.registerFont(TTFont('Jd-Lcd-Rounded', 'JdLcdRoundedRegular-vXwE.ttf'))
pdfmetrics.registerFont(TTFont('ticketing.regular', 'ticketing.regular.ttf'))
pdfmetrics.registerFont(TTFont('allerta_medium', 'allerta_medium.ttf'))
# pdfmetrics.registerFont(TTFont('Romanesque_Serif', 'Romanesque Serif.ttf'))


LOGO = "static/assests/img/logo-anderquin.png"

class Background(Flowable):
    def __init__(self, width=200, height=100, obj=None):
        self.width = width
        self.height = height
        self.obj = obj

    def wrap(self, *args):
        """Provee el tamaño del área de dibujo"""
        return (self.width, self.height)

    def draw(self):
        canvas = self.canv
        canvas.saveState()
        canvas.setLineWidth(1)
        # canvas.setFillColor(black)
        canvas.setFillColor(Color(0, 0, 0, alpha=0.4))

        # canvas.drawImage(LOGO, 200, -200, mask='auto', width=150, height=150)
        # canvas.setStrokeGray(0.1)

        # canvas.drawImage(firma, 195, -444, mask='auto', width=150, height=140)
        canvas.setFont('Narrow', 12)
        canvas.setFont('Square', 9)
        canvas.setFillColor(white)
        canvas.restoreState()


def qr_code(table):
    # generate and rescale QR
    qr_code = qr.QrCodeWidget(table)
    bounds = qr_code.getBounds()
    width = bounds[2] - bounds[0]
    height = bounds[3] - bounds[1]
    drawing = Drawing(
        3.5 * cm, 3.5 * cm, transform=[3.5 * cm / width, 0, 0, 3.5 * cm / height, 0, 0])
    drawing.add(qr_code)

    return drawing


def print_pdf_bill_finances(request, pk=None):
    _a4 = (8.3 * inch, 11.7 * inch)
    ml = 0.25 * inch
    mr = 0.25 * inch
    ms = 0.25 * inch
    mi = 0.25 * inch

    _bts = 8.3 * inch - 0.25 * inch - 0.25 * inch

    I = Image(LOGO)
    I.drawHeight = 3.60 * inch / 2.9
    I.drawWidth = 3.9 * inch / 2.9

    bill_obj = BillPurchase.objects.get(id=pk)

    tbl1_col__2 = [
        [Paragraph('INDUSTRIAS ANDERQUIN EIRL', styles["Justify_Newgot_title"])],
        [Paragraph('JR. CARABAYA NRO. 443', styles['Normal'])],
        [Paragraph('JULIACA - SAN ROMAN - PUNO', styles['Normal'])],
        ['Telefono: ' + str('999999999')],
        ['Correo: ' + str('email@anderquin.com')],
    ]
    col_2 = Table(tbl1_col__2)
    style_table_col_2 = [
        # ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]
    col_2.setStyle(TableStyle(style_table_col_2))

    tbl1_col__3 = [
        [Paragraph('RUC 20604193053', styles["Center_Newgot_title"])],
        [Paragraph('FACTURA ' + ' ELECTRÓNICA', styles["Center_Newgot_title"])],
        [Paragraph(str(bill_obj.serial.upper()) + '-' + str(bill_obj.correlative), styles["Center_Newgot_title"])],
    ]
    col_3 = Table(tbl1_col__3, colWidths=[_bts * 28 / 100])
    style_table_col_3 = [
        # ('GRID', (0, 0), (-1, -1), 0.9, colors.red),  # all columns
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # all columns
        ('ALIGNMENT', (0, 0), (0, -1), 'CENTER'),  # first column
        ('SPAN', (0, 0), (0, 0)),  # first row
    ]
    col_3.setStyle(TableStyle(style_table_col_3))

    colwidth_table_1 = [_bts * 20 / 100, _bts * 50 / 100, _bts * 30 / 100]

    style_table_header = [
        # ('GRID', (0, 0), (-1, -1), 0.9, colors.blue),  # all columns
        ('VALIGN', (1, 0), (1, 0), 'TOP'),  # all columns
        ('VALIGN', (2, 0), (2, 0), 'MIDDLE'),  # all columns
        ('ALIGNMENT', (0, 0), (0, 0), 'CENTER'),
    ]

    _tbl_header = Table(
        [(I, col_2, col_3)],
        colWidths=colwidth_table_1)
    _tbl_header.setStyle(TableStyle(style_table_header))

    # ---------------------------------Datos Cliente----------------------------#
    # client_id = bill_obj.client.id
    # client_obj = Client.objects.get(id=client_id)
    # type_client = client_obj.clienttype_set.first().document_type.id
    # info_document = client_obj.clienttype_set.first().document_number
    # telephone = client_obj.phone
    # email = client_obj.email
    info_address = ''
    # if telephone is None:
    #     telephone = '-'
    # if email is None:
    #     email = '-'
    # if type_client == '06':
    #     info_address = client_obj.clientaddress_set.last().address

    client_ruc = '20604193053'
    client_name = 'INDUSTRIAS ANDERQUIN EIRL'
    client_address = 'Jr. Carabaya No. 443 (Frente a la Plaza Manco Cápac) - JULIACA - JULIACA'

    tbl2_col1 = [
        ['Fecha de emisión', Paragraph(': ' + str(bill_obj.register_date.strftime("%d/%m/%Y")), styles['Left_Square']),
         'Lugar: ', Paragraph('JULIACA', styles['Left_Square'])],
        ['Señor(es)', Paragraph(': ' + str(client_name), styles['Left_Square']), '', ''],
        ['Direccion', Paragraph(': ' + str(client_address.upper()), styles['Left_Square']), '', ''],
        ['Doc. Identidad', Paragraph(': ' + str('RUC: ' + client_ruc), styles['Left_Square']), 'Moneda:',
         Paragraph('SOL', styles['Left_Square'])],
        # ['Telefono          :', Paragraph(str(telephone), styles['Left_Square'])],
        # ['Correo             :', Paragraph(str(email), styles['Left_Square'])],
        # ['Des. Pago : ', Paragraph(str(bill_obj), styles['Left_Square'])],
    ]
    header_description = Table(tbl2_col1,
                               colWidths=[_bts * 15 / 100, _bts * 50 / 100, _bts * 15 / 100, _bts * 20 / 100])
    style_table2_col1 = [
        # ('GRID', (0, 0), (-1, -1), 0.9, colors.blue),  # all columns
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # all columns
        ('ALIGNMENT', (0, 0), (0, -1), 'LEFT'),  # first column
        ('ALIGNMENT', (2, 0), (2, 0), 'RIGHT'),  # first column
        ('ALIGNMENT', (2, 3), (2, 3), 'RIGHT'),  # first column
        ('SPAN', (1, 1), (3, 1)),
        ('SPAN', (1, 2), (3, 2)),
        # ('BOX', (0, 0), (-1, -1), 0.9, colors.black),
        # ('BACKGROUND', (2, 0), (2, 0), colors.red),
        # ('BACKGROUND', (2, 3), (2, 3), colors.red),
        # ('LEFTPADDING', (0, 0), (0, -1), 13),  # first column
    ]
    header_description.setStyle(TableStyle(style_table2_col1))

    array_oc = []
    str_array_oc = ''
    # for p in bill_obj.purchase.all():
    #     if p.correlative:
    #         oc = p.correlative
    #     else:
    #         oc = int(p.bill_number[-5:])
    #     array_oc.append(oc)
    #     str_array_oc = str(array_oc).replace('[', '').replace(']', '')

    tbl3_col1 = [
        [Paragraph('Fecha de Pedido', styles['Center_Square']), Paragraph('Nº de Pedido', styles['Center_Square']),
         Paragraph('Nº O/C', styles['Center_Square'])],
        [Paragraph(str(bill_obj.register_date.strftime("%d/%m/%Y")), styles['Center_Square']),
         Paragraph(str(bill_obj.order_number), styles['Center_Square']),
         Paragraph(str(bill_obj.purchase.bill_number), styles['Center_Square'])]
    ]
    tbl2_col_1 = Table(tbl3_col1, colWidths=[_bts * 15 / 100, _bts * 15 / 100, _bts * 30 / 100])
    style_table2_col1 = [
        ('GRID', (0, 0), (-1, -1), 0.9, colors.black),  # all columns
        ('BACKGROUND', (0, 0), (3, 0), colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # all columns
        ('ALIGNMENT', (0, 0), (-1, -1), 'CENTER'),
    ]
    tbl2_col_1.setStyle(TableStyle(style_table2_col1))

    tbl3_col2 = [
        [Paragraph('Condición de Pago', styles['Center_Square']),
         Paragraph('Fecha de Vencimiento', styles['Center_Square'])],
        [Paragraph('Factura a ' + str(bill_obj.payment_condition) + ' días', styles['Center_Square']),
         Paragraph(str(bill_obj.expiration_date.strftime("%d/%m/%Y")), styles['Center_Square'])],
    ]
    tbl2_col_2 = Table(tbl3_col2, colWidths=[_bts * 19.5 / 100])
    style_table2_col1 = [
        ('GRID', (0, 0), (-1, -1), 0.9, colors.black),  # all columns
        ('BACKGROUND', (0, 0), (2, 0), colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # all columns
        ('ALIGNMENT', (0, 0), (-1, -1), 'CENTER'),
    ]
    tbl2_col_2.setStyle(TableStyle(style_table2_col1))

    _tbl_description_2 = [
        [tbl2_col_1, tbl2_col_2],
    ]
    description_2 = Table(_tbl_description_2, colWidths=[_bts * 60 / 100, _bts * 40 / 100])
    style_table_header = [
        # ('GRID', (0, 0), (-1, -1), 0.9, colors.blue),  # all columns
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # all columns
        ('ALIGNMENT', (0, 0), (0, -1), 'CENTER'),  # first column
        # ('SPAN', (0, 0), (0, 0)),  # first row
        # ('GRID', (0, 0), (-1, -1), 2, colors.lightgrey),
        # ('GRID', (0, 0), (0, 1), 3.5, colors.red)
    ]
    description_2.setStyle(TableStyle(style_table_header))

    style_table_header_detail = [
        ('FONTNAME', (0, 0), (-1, -1), 'Newgot'),  # all columns
        ('GRID', (0, 0), (-1, -1), 1, colors.darkgray),  # all columns
        ('BACKGROUND', (0, 0), (-1, -1), colors.darkgray),  # all columns
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 10),  # all columns
        ('LEFTPADDING', (0, 0), (0, -1), 10),  # first column
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),  # all columns
        ('RIGHTPADDING', (1, 0), (1, -1), 10),  # second column
        ('ALIGNMENT', (0, 0), (-1, -1), 'CENTER'),  # all column
        ('ALIGNMENT', (2, 0), (2, -1), 'LEFT'),  # second column
    ]
    width_table = [_bts * 8 / 100, _bts * 8 / 100, _bts * 7 / 100, _bts * 41 / 100, _bts * 10 / 100, _bts * 8 / 100,
                   _bts * 8 / 100, _bts * 10 / 100]
    header_detail = Table(
        [('Código', 'Cantidad', 'U. Venta', 'Descripción', 'Precio U.', 'V. Lista', 'Dsctos(%)', 'V. Venta Total')],
        colWidths=width_table)
    header_detail.setStyle(TableStyle(style_table_header_detail))

    style_table_detail = [
        ('FONTNAME', (0, 0), (-1, -1), 'Square'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.darkgray),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('LEFTPADDING', (0, 0), (0, -1), 10),  # first column
        ('ALIGNMENT', (0, 0), (-1, -1), 'CENTER'),  # all column
        ('ALIGNMENT', (2, 0), (2, -1), 'CENTER'),  # three column
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # first column
        ('RIGHTPADDING', (3, 0), (3, -1), 10),  # first column
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),  # all columns
        # ('BACKGROUND', (1, 0), (1, -1),  colors.green),  # four column
    ]
    detail_rows = []
    _total = 0

    for d in bill_obj.purchase.purchasedetail_set.all():
        _product = Paragraph(str(d.product.name), styles["Justify_Square"])
        _detail_total = '{:,}'.format(round(d.quantity * d.price_unit, 2))
        detail_rows.append(
            (str(d.product.code), str(decimal.Decimal(round(d.quantity, 2))), str(d.unit.description), _product,
             str(decimal.Decimal(round(d.price_unit, 2))), str(0.00), str(0.00), _detail_total))
        _total = _total + d.quantity * d.price_unit
    detail_body = Table(detail_rows,
                        colWidths=width_table)
    detail_body.setStyle(TableStyle(style_table_detail))

    sub_total = decimal.Decimal(_total) / decimal.Decimal(1.18)
    igv = decimal.Decimal(_total) - sub_total

    table_bank = [
        [Paragraph('BANCO', styles['Center_Newgot_1']),
         Paragraph('MONEDA', styles['Center_Newgot_1']),
         Paragraph('CODIGO DE CUENTA CORRIENTE', styles['Center_Newgot_1']),
         Paragraph('CODIGO DE CUENTA INTERBANCARIO', styles['Center_Newgot_1'])],

        [Paragraph('CUENTAS BCP', styles['Center_Newgot_1']),
         Paragraph('SOLES', styles['Center-text']), Paragraph('215-2023417-0-71', styles['Center-text']),
         Paragraph('002-215-002023417071-28', styles['Center-text'])],

        [Paragraph('CUENTAS BCP', styles['Center_Newgot_1']),
         Paragraph('SOLES', styles['Center-text']), Paragraph('215-9844079-0-56', styles['Center-text']),
         Paragraph('002-215-009844079056-20', styles['Center-text'])],
    ]
    t_bank = Table(table_bank, colWidths=[_bts * 7 / 100, _bts * 7 / 100, _bts * 24 / 100, _bts * 24 / 100])
    style_bank = [
        ('SPAN', (0, 1), (0, 2)),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('ALIGNMENT', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.darkgray),
    ]
    t_bank.setStyle(TableStyle(style_bank))

    total_col1 = [
        [t_bank],
        [Paragraph('OBSERVACION: ' + ' ', styles["Justify_Newgot"])],
        [Paragraph('SON: ' + numero_a_moneda(round(_total, 2), ),
                   styles["Justify_Newgot"])],
    ]
    total_col_1 = Table(total_col1, colWidths=[_bts * 63 / 100])
    style_table_col1 = [
        ('RIGHTPADDING', (0, 0), (-1, -1), 20),
        ('ALIGNMENT', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('LEFTPADDING', (0, 0), (-1, -1), -5),
        # ('GRID', (0, 0), (-1, -1), 0.5, colors.blue),
    ]
    total_col_1.setStyle(TableStyle(style_table_col1))

    money = 'S/.'
    _text = 'DESCUENTO'
    _discount = 0.00

    total_col2 = [
        [Paragraph('GRAVADA', styles["Justify_Newgot"]),
         Paragraph(money + ' ' + str('{:,}'.format(round(sub_total, 2))), styles["Right_Newgot"])],
        [Paragraph(_text, styles["Justify_Newgot"]),
         Paragraph(money + ' ' + str(round(_discount, 2)), styles["Right_Newgot"])],
        [Paragraph('I.G.V.(18.00 %)', styles["Justify_Newgot"]),
         Paragraph(money + ' ' + str('{:,}'.format(round(igv, 2))), styles["Right_Newgot"])],
        [Paragraph('TOTAL', styles["Justify_Newgot"]),
         Paragraph(money + ' ' + str('{:,}'.format(round(sub_total + igv, 2))), styles["Right_Newgot"])],
    ]
    total_col_2 = Table(total_col2, colWidths=[_bts * 14 / 100, _bts * 19 / 100])

    style_table_col2 = [
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('ALIGNMENT', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.darkgray),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]
    total_col_2.setStyle(TableStyle(style_table_col2))

    total_ = [
        [total_col_1, total_col_2],
    ]
    total_page = Table(total_, colWidths=[_bts * 65 / 100, _bts * 35 / 100])
    style_table_page = [
        ('ALIGNMENT', (0, 0), (-1, -1), 'LEFT'),  # three column
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # first column
        # ('GRID', (0, 0), (-1, -1), 0.5, colors.red),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]
    total_page.setStyle(TableStyle(style_table_page))

    _dictionary = []
    _dictionary.append(_tbl_header)
    _dictionary.append(OutputInvoiceGuide(count_row=1))
    _dictionary.append(Spacer(1, 5))
    _dictionary.append(header_description)
    _dictionary.append(Spacer(1, 5))
    _dictionary.append(description_2)
    _dictionary.append(Spacer(1, 10))
    _dictionary.append(header_detail)
    _dictionary.append(detail_body)
    _dictionary.append(Spacer(1, 10))
    _dictionary.append(total_page)

    buff = io.BytesIO()
    doc = SimpleDocTemplate(buff,
                            pagesize=(8.3 * inch, 11.7 * inch),
                            rightMargin=mr,
                            leftMargin=ml,
                            topMargin=ms,
                            bottomMargin=mi,
                            title='Factura'
                            )
    doc.build(_dictionary)

    response = HttpResponse(content_type='application/pdf')
    # response['Content-Disposition'] = 'attachment; filename="[{}].pdf"'.format(str('Factura:') + str(bill_obj.serial) + str(bill_obj.correlative))
    response.write(buff.getvalue())

    buff.close()
    return response


class OutputInvoiceGuide(Flowable):
    def __init__(self, width=200, height=3, count_row=None):
        self.width = width
        self.height = height
        self.count_row = count_row

    def wrap(self, *args):
        """Provee el tamaño del área de dibujo"""
        return (self.width, self.height)

    def draw(self):
        canvas = self.canv  # El atributo que permite dibujar en canvas

        canvas.saveState()
        canvas.setLineWidth(1)
        canvas.setFillColor(Color(0, 0, 0, alpha=0.1))
        # canvas.setFont('Newgot', 30)
        # canvas.setFillColorRGB(0.5, 0.5, 0.5)
        # canvas.roundRect(395, 8, 155, 80, 10, stroke=1, fill=0)
        # canvas.roundRect(0, -105, 550, 105, 10, stroke=1, fill=0)
        canvas.roundRect(386, 8, 169, 80, 10, stroke=1, fill=1)
        # canvas.roundRect(-7, -140, 563, 140, 10, stroke=1, fill=0)
        # canvas.roundRect(0, -(d + 110), 550, d, 10, stroke=1, fill=0)
        # canvas.roundRect(left, bottom, width, height, radius, activa borde, activa color fondo)
        canvas.restoreState()


def get_bills_data(user_id, supplier_id, payment_status='all', start_date=None, end_date=None):
    """Función auxiliar para obtener datos de facturas"""
    user_obj = User.objects.get(id=user_id)
    supplier_obj = Supplier.objects.get(id=supplier_id)
    
    bill_set = Bill.objects.filter(supplier=supplier_obj).select_related('supplier')
    
    # Filtrar por rango de fechas si se proporcionan
    if start_date and end_date:
        bill_set = bill_set.filter(register_date__range=[start_date, end_date])
    elif start_date:
        bill_set = bill_set.filter(register_date__gte=start_date)
    elif end_date:
        bill_set = bill_set.filter(register_date__lte=end_date)
    
    # Filtrar por status_pay según el estado de pago solicitado
    if payment_status == 'paid':
        bill_set = bill_set.filter(status_pay='C')  # COMPLETADA
    elif payment_status == 'pending':
        bill_set = bill_set.filter(status_pay='P')  # PENDIENTE
    
    bill_set = bill_set.order_by('register_date')
    
    return bill_set, supplier_obj


def print_pdf_purchases_report(request):
    """Generar PDF del reporte de estado de cuentas"""
    if request.method == 'GET':
        user_id = request.user.id
        supplier_id = request.GET.get('supplier_id')
        payment_status = request.GET.get('payment_status', 'all')
        start_date = request.GET.get('start_date', None)
        end_date = request.GET.get('end_date', None)
        
        if not supplier_id:
            return HttpResponse('Error: Proveedor no especificado', status=400)
        
        bill_set, supplier_obj = get_bills_data(user_id, supplier_id, payment_status, start_date, end_date)
        
        # Configuración del PDF
        _a4 = (8.3 * inch, 11.7 * inch)
        ml = 0.25 * inch
        mr = 0.25 * inch
        ms = 0.25 * inch
        mi = 0.25 * inch
        _bts = 8.3 * inch - 0.25 * inch - 0.25 * inch
        
        # Logo
        I = Image(LOGO)
        I.drawHeight = 2 * inch
        I.drawWidth = 2.5 * inch
        
        # Encabezado
        title = "REPORTE DE ESTADO DE CUENTAS"
        if payment_status == 'paid':
            title = "REPORTE DE FACTURAS PAGADAS"
        elif payment_status == 'pending':
            title = "REPORTE DE FACTURAS PENDIENTES"
        
        # Formatear rango de fechas si se proporciona
        date_range_text = 'Fecha: ' + datetime.now().strftime("%d/%m/%Y")
        if start_date and end_date:
            date_range_text = f'Rango de fechas: {start_date} al {end_date}'
        elif start_date:
            date_range_text = f'Desde: {start_date}'
        elif end_date:
            date_range_text = f'Hasta: {end_date}'
        
        tbl_header = [
            [Paragraph('INDUSTRIAS ANDERQUIN EIRL', styles["Justify_Newgot_title"])],
            [Paragraph('JR. CARABAYA NRO. 443 - JULIACA', styles['Normal'])],
            [Paragraph('RUC: 20604193053', styles['Normal'])],
            [Paragraph(title, styles["Center_Newgot_title_small"])],
            [Paragraph('Proveedor: ' + supplier_obj.business_name, styles['Normal'])],
            [Paragraph(date_range_text, styles['Normal'])],
        ]
        
        header_table = Table(tbl_header)
        header_table.setStyle(TableStyle([
            ('ALIGNMENT', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 3), (-1, 3), colors.lightgrey),  # Fondo gris claro para el título
            ('TEXTCOLOR', (0, 3), (-1, 3), colors.black),  # Texto negro para el título
            ('FONTSIZE', (0, 3), (-1, 3), 12),  # Reducir tamaño de fuente del título
            ('PADDING', (0, 3), (-1, 3), 10),  # Padding para el título
        ]))
        
        # Encabezados de columnas según la imagen
        headers = ['FECHA EMISION', 'Nº DE DOCUMENTO', 'IMPORTE FAC', 'FECHA DEPOSITO', 
                  'IMPORTE DEPOSITADO', 'ESTADO', 'SALDO', 'PROVEEDOR']
        
        header_row = [Paragraph(h, styles['Center_Square']) for h in headers]
        
        # Datos
        data_rows = []
        sum_total_deposited = 0
        
        for bill in bill_set:
            # Obtener el monto pagado, asegurándose de que sea un valor válido
            repay_loan_value = bill.repay_loan()
            if repay_loan_value is None:
                repay_loan_value = 0
            repay_loan = decimal.Decimal(repay_loan_value)
            bill_total = decimal.Decimal(bill.bill_total_total)
            missing_payment = round(bill_total, 2) - round(repay_loan, 2)
            
            # Obtener fecha del último depósito - mejorado para facturas pagadas
            # Primero intentar obtener el pago más reciente con operation_date
            last_payment_with_date = bill.loanpayment_set.filter(
                type='C', 
                operation_date__isnull=False
            ).order_by('-operation_date').first()
            
            # Si no hay uno con operation_date, buscar el más reciente por create_at
            if not last_payment_with_date:
                last_payment_with_date = bill.loanpayment_set.filter(
                    type='C'
                ).order_by('-create_at').first()
            
            deposit_date = '-'
            if last_payment_with_date:
                if last_payment_with_date.operation_date:
                    deposit_date = last_payment_with_date.operation_date.strftime("%d/%m/%Y")
                elif last_payment_with_date.create_at:
                    deposit_date = last_payment_with_date.create_at.date().strftime("%d/%m/%Y")
            
            # Estado
            estado = 'CANCELADO' if missing_payment == 0 else 'PENDIENTE'
            
            # Saldo (vacío si está cancelado)
            saldo = '' if missing_payment == 0 else f"S/ {missing_payment:,.2f}"
            
            row = [
                Paragraph(bill.register_date.strftime("%d/%m/%Y"), styles['Center_Square']),
                Paragraph(f"{bill.serial}-{str(bill.correlative).zfill(7)}", styles['Center_Square']),
                Paragraph(f"S/ {bill_total:,.2f}", styles['Center_Square']),
                Paragraph(deposit_date, styles['Center_Square']),
                Paragraph(f"S/ {repay_loan:,.2f}", styles['Center_Square']),
                Paragraph(estado, styles['Center_Square']),
                Paragraph(saldo, styles['Center_Square']),
                Paragraph(supplier_obj.business_name, styles['Center_Square']),
            ]
            data_rows.append(row)
            
            sum_total_deposited += repay_loan
        
        # Tabla de datos
        all_data = [header_row] + data_rows
        # Usar todo el ancho disponible de A4
        available_width = _bts - 20
        col_widths = [available_width * 0.12, available_width * 0.14, available_width * 0.12, 
                     available_width * 0.12, available_width * 0.14, available_width * 0.12, 
                     available_width * 0.12, available_width * 0.12]
        
        data_table = Table(all_data, colWidths=col_widths)
        data_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#E1D5E7')),  # Fondo morado claro para el encabezado
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGNMENT', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Square-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            # Alternar colores de fila
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, HexColor('#FFFACD')]),
        ]))
        
        # Fila de TOTAL
        totals_data = [
            ['', '', '', 'TOTAL', f"S/ {sum_total_deposited:,.2f}", '', '', '']
        ]
        
        totals_table = Table(totals_data, colWidths=col_widths)
        totals_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#FFFF00')),  # Fondo amarillo para el total
            ('FONTNAME', (0, 0), (-1, 0), 'Square-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGNMENT', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        
        # Construir PDF
        buff = io.BytesIO()
        doc = SimpleDocTemplate(buff, pagesize=A4, rightMargin=10, leftMargin=10, 
                              topMargin=15, bottomMargin=15, title='Reporte Estado de Cuentas')
        
        elements = []
        elements.append(header_table)
        elements.append(Spacer(1, 20))
        elements.append(data_table)
        elements.append(Spacer(1, 10))
        elements.append(totals_table)
        
        doc.build(elements)
        
        response = HttpResponse(content_type='application/pdf')
        filename = f"reporte_estado_cuentas_{supplier_obj.business_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(buff.getvalue())
        
        buff.close()
        return response
    
    return HttpResponse('Método no permitido', status=405)


def export_excel_purchases_report(request):
    """Generar Excel del reporte de estado de cuentas"""
    if request.method == 'GET':
        user_id = request.user.id
        supplier_id = request.GET.get('supplier_id')
        payment_status = request.GET.get('payment_status', 'all')
        start_date = request.GET.get('start_date', None)
        end_date = request.GET.get('end_date', None)
        
        if not supplier_id:
            return HttpResponse('Error: Proveedor no especificado', status=400)
        
        bill_set, supplier_obj = get_bills_data(user_id, supplier_id, payment_status, start_date, end_date)
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Configurar estilos
        title_font = Font(name='Arial', size=16, bold=True)
        header_font = Font(name='Arial', size=10, bold=True)
        data_font = Font(name='Arial', size=9)
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        title_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        title = "REPORTE DE ESTADO DE CUENTAS"
        if payment_status == 'paid':
            title = "REPORTE DE FACTURAS PAGADAS"
        elif payment_status == 'pending':
            title = "REPORTE DE FACTURAS PENDIENTES"
        
        ws.merge_cells('A1:H1')
        ws['A1'] = 'INDUSTRIAS ANDERQUIN EIRL'
        ws['A1'].font = title_font
        ws['A1'].alignment = center_alignment
        ws['A1'].fill = title_fill
        
        ws.merge_cells('A2:H2')
        ws['A2'] = 'JR. CARABAYA NRO. 443 - JULIACA - RUC: 20604193053'
        ws['A2'].font = data_font
        ws['A2'].alignment = center_alignment
        
        ws.merge_cells('A3:H3')
        ws['A3'] = title
        ws['A3'].font = Font(name='Arial', size=14, bold=True)
        ws['A3'].alignment = center_alignment
        ws['A3'].fill = title_fill
        
        ws.merge_cells('A4:H4')
        ws['A4'] = f'Proveedor: {supplier_obj.business_name}'
        ws['A4'].font = data_font
        ws['A4'].alignment = center_alignment
        
        # Formatear rango de fechas si se proporciona
        date_range_text = f'Fecha: {datetime.now().strftime("%d/%m/%Y")}'
        if start_date and end_date:
            date_range_text = f'Rango de fechas: {start_date} al {end_date}'
        elif start_date:
            date_range_text = f'Desde: {start_date}'
        elif end_date:
            date_range_text = f'Hasta: {end_date}'
        
        ws.merge_cells('A5:H5')
        ws['A5'] = date_range_text
        ws['A5'].font = data_font
        ws['A5'].alignment = center_alignment
        
        # Encabezados según la imagen con formato de dos líneas
        headers = [
            'FECHA\nEMISION',           # Columna 1
            'Nº DE\nDOCUMENTO',         # Columna 2
            'IMPORTE\nFAC',             # Columna 3
            'FECHA\nDEPOSITO',          # Columna 4
            'IMPORTE\nDEPOSITADO',      # Columna 5
            'ESTADO',                   # Columna 6
            'SALDO',                    # Columna 7
            'PROVEEDOR'                 # Columna 8
        ]
        
        # Color morado claro para el encabezado
        header_fill = PatternFill(start_color='E1D5E7', end_color='E1D5E7', fill_type='solid')
        
        # Alineación con wrap_text para permitir múltiples líneas
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Datos
        row = 8
        sum_total_deposited = 0
        
        for bill in bill_set:
            # Obtener el monto pagado, asegurándose de que sea un valor válido
            repay_loan_value = bill.repay_loan()
            if repay_loan_value is None:
                repay_loan_value = 0
            repay_loan = decimal.Decimal(repay_loan_value)
            bill_total = decimal.Decimal(bill.bill_total_total)
            missing_payment = round(bill_total, 2) - round(repay_loan, 2)
            
            # Obtener fecha del último depósito - mejorado para facturas pagadas
            # Primero intentar obtener el pago más reciente con operation_date
            last_payment_with_date = bill.loanpayment_set.filter(
                type='C', 
                operation_date__isnull=False
            ).order_by('-operation_date').first()
            
            # Si no hay uno con operation_date, buscar el más reciente por create_at
            if not last_payment_with_date:
                last_payment_with_date = bill.loanpayment_set.filter(
                    type='C'
                ).order_by('-create_at').first()
            
            deposit_date = '-'
            if last_payment_with_date:
                if last_payment_with_date.operation_date:
                    deposit_date = last_payment_with_date.operation_date.strftime("%d/%m/%Y")
                elif last_payment_with_date.create_at:
                    deposit_date = last_payment_with_date.create_at.date().strftime("%d/%m/%Y")
            
            # Estado
            estado = 'CANCELADO' if missing_payment == 0 else 'PENDIENTE'
            
            # Saldo (vacío si está cancelado)
            saldo = '' if missing_payment == 0 else float(missing_payment)
            
            # Datos de la fila
            ws.cell(row=row, column=1, value=bill.register_date.strftime("%d/%m/%Y")).border = thin_border
            ws.cell(row=row, column=2, value=f"{bill.serial}-{str(bill.correlative).zfill(7)}").border = thin_border
            ws.cell(row=row, column=3, value=float(bill_total)).border = thin_border
            ws.cell(row=row, column=4, value=deposit_date).border = thin_border
            ws.cell(row=row, column=5, value=float(repay_loan)).border = thin_border
            ws.cell(row=row, column=6, value=estado).border = thin_border
            ws.cell(row=row, column=7, value=saldo).border = thin_border
            ws.cell(row=row, column=8, value=supplier_obj.business_name).border = thin_border
            
            # Aplicar formato contable personalizado con S/ a números
            # Formato: positivo;negativo;cero;texto
            custom_format = '_("S/"* #,##0.00_);_("S/"* (#,##0.00);_("S/"* "-"??_);_(@_)'
            ws.cell(row=row, column=3).number_format = custom_format
            if saldo != '':
                ws.cell(row=row, column=7).number_format = custom_format
            ws.cell(row=row, column=5).number_format = custom_format
            
            # Aplicar fuente y alineación a toda la fila
            for col in range(1, 9):
                ws.cell(row=row, column=col).font = data_font
                ws.cell(row=row, column=col).alignment = center_alignment
            
            # Alternar colores de fila (blanco y amarillo claro)
            # La fila 8 es par (índice 0), así que empezamos con amarillo claro
            if (row - 8) % 2 == 0:
                fill_color = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
            else:
                fill_color = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = fill_color
            
            sum_total_deposited += repay_loan
            
            row += 1
        
        # Fila de TOTAL con fondo amarillo
        total_row = row + 1
        total_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        ws.cell(row=total_row, column=1, value='').border = thin_border
        ws.cell(row=total_row, column=2, value='').border = thin_border
        ws.cell(row=total_row, column=3, value='').border = thin_border
        ws.cell(row=total_row, column=4, value='TOTAL').border = thin_border
        ws.cell(row=total_row, column=4).font = header_font
        ws.cell(row=total_row, column=5, value=float(sum_total_deposited)).border = thin_border
        ws.cell(row=total_row, column=5).font = header_font
        ws.cell(row=total_row, column=5).number_format = '_("S/"* #,##0.00_);_("S/"* (#,##0.00);_("S/"* "-"??_);_(@_)'
        ws.cell(row=total_row, column=6, value='').border = thin_border
        ws.cell(row=total_row, column=7, value='').border = thin_border
        ws.cell(row=total_row, column=8, value='').border = thin_border
        
        # Aplicar fondo amarillo a toda la fila de total
        for col in range(1, 9):
            ws.cell(row=total_row, column=col).fill = total_fill
            ws.cell(row=total_row, column=col).alignment = center_alignment
        
        # Ajustar ancho de columnas y altura de fila de encabezado
        column_widths = [12, 15, 12, 12, 15, 12, 12, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Ajustar altura de la fila de encabezado para que se vean las dos líneas
        ws.row_dimensions[7].height = 30
        
        # Guardar
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"reporte_estado_cuentas_{supplier_obj.business_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    return HttpResponse('Método no permitido', status=405)
